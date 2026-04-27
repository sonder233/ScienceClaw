from __future__ import annotations

import json
import re
from typing import Any, Dict, Iterable, List, Optional

from backend.rpa.playwright_security import get_chromium_launch_kwargs, get_context_kwargs
from backend.rpa.upload_source import (
    default_upload_source_from_staging,
    download_result_key,
    render_file_source_expr,
    render_legacy_input_files_expr,
    rpa_asset_helper_lines,
    upload_source_uses_asset_helper,
)

from .trace_models import RPAAcceptedTrace, RPATraceType


class TraceSkillCompiler:
    def generate_script(
        self,
        traces: Iterable[RPAAcceptedTrace],
        params: Optional[Dict[str, Any]] = None,
        *,
        is_local: bool = False,
        test_mode: bool = False,
    ) -> str:
        self._compiled_output_keys: Dict[int, str] = {}
        self._params = params or {}
        self._use_upload_sources = is_local
        self._param_lookup = self._build_param_lookup(self._params)
        self._param_cursors: Dict[str, int] = {}
        trace_list = list(traces)
        execute_skill_func = "\n".join(self._render_execute_skill(trace_list))
        return _runner_template(is_local).format(
            execute_skill_func=execute_skill_func,
            launch_kwargs=repr(get_chromium_launch_kwargs(headless=False)),
            context_kwargs=repr(get_context_kwargs()),
        )

    def _render_execute_skill(self, traces: List[RPAAcceptedTrace]) -> List[str]:
        lines = [
            "",
            "def _resolve_result_ref(results, ref):",
            "    current = results",
            "    for segment in str(ref).split('.'):",
            "        if isinstance(current, dict) and segment in current:",
            "            current = current[segment]",
            "            continue",
            "        if isinstance(current, list) and segment.isdigit():",
            "            current = current[int(segment)]",
            "            continue",
            "        raise KeyError(ref)",
            "    return current",
            "",
            "def _resolve_first_result_ref(results, refs):",
            "    last_error = None",
            "    for ref in refs:",
            "        try:",
            "            return _resolve_result_ref(results, ref)",
            "        except KeyError as exc:",
            "            last_error = exc",
            "    raise last_error or KeyError(refs[0] if refs else '')",
            "",
            "def _validate_non_empty_records(key, value):",
            "    if not isinstance(value, list) or not value:",
            "        raise RuntimeError(f'AI trace output {key} is empty')",
            "",
            "def _trace_page_url(page):",
            "    try:",
            "        return str(getattr(page, 'url', '') or '')",
            "    except Exception:",
            "        return ''",
            "",
            "def _trace_emit(logger, event, index, description, page, started_at=None, error=None):",
            "    if not callable(logger):",
            "        return",
            "    prefix = {'START': 'TRACE_START', 'DONE': 'TRACE_DONE', 'ERROR': 'TRACE_ERROR'}.get(event, f'TRACE_{event}')",
            "    parts = [f'{prefix} {index}: {description}']",
            "    if started_at is not None:",
            "        parts.append(f'duration_ms={(time.perf_counter() - started_at) * 1000:.1f}')",
            "    page_url = _trace_page_url(page)",
            "    if page_url:",
            "        parts.append(f'url={page_url}')",
            "    if error is not None:",
            "        message = str(error).replace('\\n', ' ')[:300]",
            "        parts.append(f'error={type(error).__name__}: {message}')",
            "    try:",
            "        logger(' | '.join(parts))",
            "    except Exception:",
            "        pass",
            "",
            "def _trace_start(logger, index, description, page):",
            "    started_at = time.perf_counter()",
            "    _trace_emit(logger, 'START', index, description, page)",
            "    return started_at",
            "",
            "def _trace_done(logger, index, description, page, started_at):",
            "    _trace_emit(logger, 'DONE', index, description, page, started_at)",
            "",
            "def _trace_error(logger, index, description, page, started_at, error):",
            "    _trace_emit(logger, 'ERROR', index, description, page, started_at, error)",
            "",
            "def _abs_github_url(href):",
            "    if not href:",
            "        return ''",
            "    if href.startswith(('http://', 'https://')):",
            "        return href",
            "    return 'https://github.com' + href",
            "",
            "def _github_repo_base(url):",
            "    text = str(url or '').split('?', 1)[0].rstrip('/')",
            "    match = re.match(r'(https://github\\.com/[^/]+/[^/]+)', text)",
            "    return match.group(1) if match else ''",
            "",
            "def _normalize_runtime_ai_payload(payload, page_url=''):",
            "    if isinstance(payload, dict) and len(payload) == 1:",
            "        only_value = next(iter(payload.values()))",
            "        if isinstance(only_value, dict):",
            "            payload = only_value",
            "    if isinstance(payload, str):",
            "        payload = {'value': payload}",
            "    if not isinstance(payload, dict):",
            "        payload = {'value': payload}",
            "    value = payload.get('value')",
            "    if 'url' not in payload and isinstance(value, str) and value.startswith(('http://', 'https://')):",
            "        payload['url'] = value",
            "    if 'url' not in payload and page_url:",
            "        payload['url'] = page_url",
            "    if 'name' not in payload and isinstance(payload.get('url'), str):",
            "        match = re.match(r'https://github\\.com/([^/]+/[^/?#]+)', payload['url'])",
            "        if match:",
            "            payload['name'] = match.group(1)",
            "    return payload",
            "",
            "async def _execute_runtime_ai_instruction(page, results, instruction, output_key):",
            "    from backend.rpa.recording_runtime_agent import RecordingRuntimeAgent",
            "    agent = RecordingRuntimeAgent()",
            "    outcome = await agent.run(page=page, instruction=instruction, runtime_results=results)",
            "    if not outcome.success:",
            "        detail = '; '.join(str(item.message) for item in outcome.diagnostics) or outcome.message",
            "        raise RuntimeError(f'Runtime semantic instruction failed: {detail}')",
            "    payload = outcome.output",
            "    if isinstance(payload, dict) and output_key in payload and isinstance(payload.get(output_key), (dict, list, str)):",
            "        payload = payload.get(output_key)",
            "    payload = _normalize_runtime_ai_payload(payload, getattr(page, 'url', ''))",
            "    if outcome.output_key and outcome.output_key not in results:",
            "        results[outcome.output_key] = payload",
            "    if output_key:",
            "        results[output_key] = payload",
            "    return payload",
            "",
            "async def execute_skill(page, **kwargs):",
            '    """Auto-generated skill from RPA trace recording."""',
            "    _results = {}",
            "    current_page = page",
            "    _trace_logger = kwargs.get('_on_log')",
        ]
        if self._use_upload_sources and any(upload_source_uses_asset_helper(_trace_upload_source(trace)) for trace in traces):
            lines.extend(rpa_asset_helper_lines("    "))
        used_output_keys: Dict[str, int] = {}
        for index, trace in enumerate(traces):
            trace_lines = self._render_trace(index, trace, traces[:index], used_output_keys)
            lines.extend(self._wrap_trace_logging(index, trace, trace_lines))
        lines.append("    return _results")
        return lines

    def _wrap_trace_logging(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        trace_lines: List[str],
    ) -> List[str]:
        description = self._trace_log_description(trace)
        wrapped = [
            "",
            f"    _trace_started_at = _trace_start(_trace_logger, {index}, {description!r}, current_page)",
            "    try:",
        ]
        for line in trace_lines:
            wrapped.append(f"    {line}" if line else "")
        wrapped.extend(
            [
                "    except Exception as _trace_exc:",
                f"        _trace_error(_trace_logger, {index}, {description!r}, current_page, _trace_started_at, _trace_exc)",
                "        raise",
                "    else:",
                f"        _trace_done(_trace_logger, {index}, {description!r}, current_page, _trace_started_at)",
            ]
        )
        return wrapped

    @staticmethod
    def _trace_log_description(trace: RPAAcceptedTrace) -> str:
        text = trace.description or trace.user_instruction or trace.action or trace.trace_type.value
        return " ".join(str(text or "").split())[:160]

    def _render_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        previous_traces: List[RPAAcceptedTrace],
        used_output_keys: Dict[str, int],
    ) -> List[str]:
        if trace.trace_type == RPATraceType.NAVIGATION:
            return self._render_navigation_trace(index, trace, previous_traces)
        if trace.trace_type == RPATraceType.DATAFLOW_FILL and trace.dataflow:
            return self._render_dataflow_fill_trace(index, trace)
        if trace.trace_type == RPATraceType.MANUAL_ACTION:
            return self._render_manual_action_trace(index, trace, previous_traces)
        if trace.trace_type == RPATraceType.DATA_CAPTURE:
            return self._render_data_capture_trace(index, trace, used_output_keys)
        if trace.trace_type == RPATraceType.FILE_TRANSFORM:
            return self._render_file_transform_trace(index, trace)
        if trace.trace_type == RPATraceType.AI_OPERATION:
            return self._render_ai_operation_trace(index, trace, previous_traces, used_output_keys)
        return ["", f"    # trace {index}: unsupported trace type {trace.trace_type.value}"]

    def _render_navigation_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        previous_traces: List[RPAAcceptedTrace],
    ) -> List[str]:
        url = trace.after_page.url or str(trace.value or "")
        dynamic = self._dynamic_url_expression(url, previous_traces)
        lines = ["", f"    # trace {index}: {trace.description or 'navigation'}"]
        if dynamic:
            lines.append(f"    _target_url = {dynamic}")
        else:
            lines.append(f"    _target_url = {url!r}")
        lines.extend(
            [
                "    await current_page.goto(_target_url, wait_until='domcontentloaded')",
                "    await current_page.wait_for_load_state('domcontentloaded')",
            ]
        )
        return lines

    def _render_manual_action_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        previous_traces: List[RPAAcceptedTrace],
    ) -> List[str]:
        action = trace.action or ""
        if action in {"navigate_click", "navigate_press"} and trace.after_page.url:
            return self._render_navigation_trace(index, trace, previous_traces)
        stable_subpage = self._manual_github_subpage_navigation(index, trace, previous_traces)
        if stable_subpage:
            return stable_subpage
        locator = self._best_locator(trace.locator_candidates)
        lines = ["", f"    # trace {index}: {trace.description or action}"]
        if not locator:
            lines.append("    # No stable locator was recorded for this manual action.")
            return lines
        expr = _locator_expression("current_page", locator)
        download_signal = _trace_download_signal(trace)
        if action in {"click", "press"} and download_signal:
            download_name = str(download_signal.get("filename") or trace.value or "file")
            result_key = str(download_signal.get("result_key") or download_result_key(download_name))
            lines.append("    async with current_page.expect_download() as _dl_info:")
            if action == "click":
                lines.append(f"        await {expr}.click()")
            else:
                lines.append(f"        await {expr}.press({str(trace.value or '')!r})")
            lines.append("    _dl = await _dl_info.value")
            self._append_download_save_lines(lines, result_key, download_name)
        elif action == "click":
            lines.append(f"    await {expr}.click()")
            lines.append("    await current_page.wait_for_timeout(500)")
        elif action == "fill":
            fill_value = self._maybe_parameterize_value(str(trace.value or ""))
            lines.append(f"    await {expr}.fill({fill_value})")
        elif action == "press":
            lines.append(f"    await {expr}.press({str(trace.value or '')!r})")
        elif action == "check":
            lines.append(f"    await {expr}.check()")
        elif action == "uncheck":
            lines.append(f"    await {expr}.uncheck()")
        elif action == "select":
            lines.append(f"    await {expr}.select_option({str(trace.value or '')!r})")
        elif action == "set_input_files":
            upload_source = _trace_upload_source(trace)
            if upload_source and self._use_upload_sources:
                file_expr = render_file_source_expr(upload_source, self._params)
            else:
                signals = trace.signals if isinstance(trace.signals, dict) else {}
                payload = signals.get("set_input_files")
                files = payload.get("files") if isinstance(payload, dict) else []
                file_expr = render_legacy_input_files_expr(files if isinstance(files, list) else [], trace.value)
                lines.append(
                    "    print('RPA_WARNING: upload trace has no configured file source; replay may fail', file=__import__('sys').stderr)"
                )
            lines.append(f"    await {expr}.set_input_files({file_expr})")
        else:
            lines.append(f"    # Unsupported manual action preserved as no-op: {action}")
        return lines

    @staticmethod
    def _append_download_save_lines(lines: List[str], result_key: str, recorded_name: str) -> None:
        lines.append("    _dl_dir = kwargs.get('_downloads_dir', '.')")
        lines.append("    import os as _os; _os.makedirs(_dl_dir, exist_ok=True)")
        lines.append(f"    _dl_filename = {json.dumps(recorded_name)} or _dl.suggested_filename")
        lines.append("    _dl_filename = _os.path.basename(str(_dl_filename)) or _dl.suggested_filename")
        lines.append("    _dl_dest = _os.path.join(_dl_dir, _dl_filename)")
        lines.append("    if _os.path.exists(_dl_dest):")
        lines.append("        _os.remove(_dl_dest)")
        lines.append("    await _dl.save_as(_dl_dest)")
        lines.append(f"    _results[{result_key!r}] = {{'filename': _dl_filename, 'path': _dl_dest}}")

    def _render_data_capture_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        used_output_keys: Dict[str, int],
    ) -> List[str]:
        locator = self._best_locator(trace.locator_candidates)
        key = self._allocate_output_key(trace, trace.output_key or f"capture_{index}", used_output_keys)
        lines = ["", f"    # trace {index}: {trace.description or 'data capture'}"]
        if locator:
            lines.append(f"    _result = await {_locator_expression('current_page', locator)}.inner_text()")
        else:
            lines.append(f"    _result = {trace.output!r}")
        lines.append(f"    _results[{key!r}] = _result")
        return lines

    def _render_file_transform_trace(self, index: int, trace: RPAAcceptedTrace) -> List[str]:
        signals = trace.signals if isinstance(trace.signals, dict) else {}
        transform = signals.get("file_transform") if isinstance(signals, dict) else {}
        transform = transform if isinstance(transform, dict) else {}
        input_source = transform.get("input") if isinstance(transform.get("input"), dict) else {}
        source_result_key = str(input_source.get("source_result_key") or transform.get("source_result_key") or "")
        file_field = str(input_source.get("file_field") or "path")
        output_result_key = str(transform.get("output_result_key") or trace.output_key or "")
        output_filename = str(transform.get("output_filename") or trace.value or "converted.xlsx")
        instruction = str(transform.get("instruction") or trace.user_instruction or trace.description or "")
        code = str(transform.get("code") or (trace.ai_execution.code if trace.ai_execution else "") or "")
        lines = ["", f"    # trace {index}: {trace.description or 'file transform'}"]
        if not source_result_key or not output_result_key or not code:
            lines.append("    # File transform trace is missing source/result/code and was skipped.")
            return lines
        lines.extend(
            [
                "    import os as _os",
                f"    _transform_input = _results[{source_result_key!r}][{file_field!r}]",
                "    _transform_dir = kwargs.get('_downloads_dir', '.')",
                "    _os.makedirs(_transform_dir, exist_ok=True)",
                f"    _transform_output = _os.path.join(_transform_dir, {output_filename!r})",
                "    _transform_ns = {'__name__': 'rpa_transform'}",
                f"    exec({code!r}, _transform_ns, _transform_ns)",
                "    _transform_fn = _transform_ns.get('transform_file')",
                "    if not callable(_transform_fn):",
                "        raise RuntimeError('file transform script has no transform_file function')",
                f"    _transform_fn(_transform_input, _transform_output, {instruction!r})",
                "    if not _os.path.exists(_transform_output):",
                "        raise RuntimeError('file transform did not create output file')",
                "    from openpyxl import load_workbook as _load_workbook",
                "    _wb = _load_workbook(_transform_output, read_only=True, data_only=True)",
                "    _wb.close()",
                f"    _results[{output_result_key!r}] = {{'filename': {output_filename!r}, 'path': _transform_output}}",
            ]
        )
        return lines

    def _render_ai_operation_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        previous_traces: List[RPAAcceptedTrace],
        used_output_keys: Dict[str, int],
    ) -> List[str]:
        instruction = f"{trace.user_instruction or ''} {trace.description or ''}".lower()
        if _looks_like_highest_star(instruction):
            return self._render_highest_star_trace(index, trace, used_output_keys)
        if _looks_like_pr_extraction(instruction, trace.output):
            return self._render_pr_extraction_trace(index, trace, previous_traces, used_output_keys)
        if _looks_like_semantic_repo_selection(instruction, trace.output):
            return self._render_semantic_repo_selection_trace(index, trace, used_output_keys)
        if trace.ai_execution and trace.ai_execution.code:
            return self._render_embedded_ai_code_trace(index, trace, previous_traces, used_output_keys)
        return ["", f"    # trace {index}: AI operation has no executable body"]

    def _render_highest_star_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        used_output_keys: Dict[str, int],
    ) -> List[str]:
        key = self._allocate_output_key(trace, trace.output_key or "selected_project", used_output_keys)
        return [
            "",
            f"    # trace {index}: generalized highest-star repository selection",
            "    rows = await current_page.locator('article.Box-row').all()",
            "    max_stars = -1",
            "    _result = None",
            "    for row in rows:",
            "        try:",
            "            star_text = (await row.locator('a[href*=\"/stargazers\"]').first.inner_text()).strip()",
            "            normalized = star_text.replace(',', '').strip().lower()",
            "            match = re.search(r'\\d+(?:\\.\\d+)?', normalized)",
            "            if not match:",
            "                continue",
            "            stars = float(match.group(0))",
            "            if 'k' in normalized:",
            "                stars *= 1000",
            "            elif 'm' in normalized:",
            "                stars *= 1000000",
            "            link = row.locator('h2 a').first",
            "            href = await link.get_attribute('href')",
            "            name = (await link.inner_text()).strip()",
            "            if href and stars > max_stars:",
            "                max_stars = stars",
            "                _result = {'name': name.replace(' ', ''), 'url': _abs_github_url(href), 'stars': int(stars)}",
            "        except Exception:",
            "            continue",
            "    if not _result:",
            "        raise RuntimeError('No repository rows with star counts were found')",
            "    await current_page.goto(_result['url'], wait_until='domcontentloaded')",
            "    await current_page.wait_for_load_state('domcontentloaded')",
            f"    _results[{key!r}] = _result",
        ]

    def _render_semantic_repo_selection_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        used_output_keys: Dict[str, int],
    ) -> List[str]:
        key = self._allocate_output_key(trace, trace.output_key or "selected_project", used_output_keys)
        return [
            "",
            f"    # trace {index}: runtime semantic repository selection",
            f"    _result = await _execute_runtime_ai_instruction(current_page, _results, {str(trace.user_instruction or trace.description or '').strip()!r}, {key!r})",
            "    if not isinstance(_result, dict) or not _result.get('url'):",
            "        raise RuntimeError('Runtime semantic selection did not produce a repository URL')",
        ]

    def _render_pr_extraction_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        previous_traces: List[RPAAcceptedTrace],
        used_output_keys: Dict[str, int],
    ) -> List[str]:
        key = self._allocate_output_key(trace, trace.output_key or "top10_prs", used_output_keys)
        allow_empty = isinstance(trace.output, list) and not trace.output
        instruction = f"{trace.user_instruction or ''} {trace.description or ''}"
        page_count = _extract_requested_page_count(instruction)
        record_limit = _extract_record_limit(instruction)
        previous_repo_expr = self._previous_repo_url_expression(previous_traces)
        lines = [
            "",
            f"    # trace {index}: generalized PR record extraction",
            f"    _page_count = {page_count}",
            f"    _record_limit = {record_limit if record_limit is not None else 'None'}",
            "    _repo_base = _github_repo_base(current_page.url)",
        ]
        if previous_repo_expr:
            lines.extend(
                [
                    "    if not _repo_base:",
                    f"        _repo_base = _github_repo_base(str({previous_repo_expr}))",
                ]
            )
        lines.extend(
            [
            "    _result = []",
            "    _seen_urls = set()",
            "    async def _collect_current_pr_rows():",
            "        rows = await current_page.locator('div.js-issue-row, div[data-testid=\"issue-row\"], div.Box-row').all()",
            "        collected = []",
            "        for row in rows:",
            "            title = ''",
            "            creator = ''",
            "            url = ''",
            "            for selector in ['a[href*=\"/pull/\"]', 'a.Link--primary', 'a[id^=\"issue_\"]', 'a.js-navigation-open']:",
            "                loc = row.locator(selector).first",
            "                if await loc.count() > 0:",
            "                    title = (await loc.inner_text()).strip()",
            "                    url = (await loc.get_attribute('href')) or ''",
            "                    if '/pull/' in url and title and not re.fullmatch(r'\\d+(\\s+comments?)?', title.lower()):",
            "                        break",
            "            for selector in ['a[data-hovercard-type=\"user\"]', 'a[href*=\"author%3A\"]', 'a[href*=\"author:\"]']:",
            "                loc = row.locator(selector).first",
            "                if await loc.count() > 0:",
            "                    creator = (await loc.inner_text()).strip()",
            "                    if creator:",
            "                        break",
            "            if title and creator:",
            "                absolute_url = _abs_github_url(url) if url else ''",
            "                if absolute_url and absolute_url in _seen_urls:",
            "                    continue",
            "                if absolute_url:",
            "                    _seen_urls.add(absolute_url)",
            "                collected.append({'title': title, 'creator': creator, 'url': absolute_url})",
            "        return collected",
            "    for _page_number in range(1, _page_count + 1):",
            "        if _repo_base:",
            "            _target_url = _repo_base + '/pulls?q=is%3Apr'",
            "            if _page_number > 1:",
            "                _target_url += f'&page={_page_number}'",
            "            await current_page.goto(_target_url, wait_until='domcontentloaded')",
            "            await current_page.wait_for_load_state('domcontentloaded')",
            "        _result.extend(await _collect_current_pr_rows())",
            "        if _record_limit is not None and len(_result) >= _record_limit:",
            "            _result = _result[:_record_limit]",
            "            break",
            "        if not _repo_base and _page_number < _page_count:",
            "            next_link = current_page.locator('a[rel=\"next\"]').first",
            "            if await next_link.count() == 0:",
            "                break",
            "            await next_link.click()",
            "            await current_page.wait_for_load_state('domcontentloaded')",
            f"    _results[{key!r}] = _result",
            ]
        )
        return lines

    @staticmethod
    def _build_param_lookup(params: Dict[str, Any]) -> Dict[str, List[tuple[str, Dict[str, Any]]]]:
        lookup: Dict[str, List[tuple[str, Dict[str, Any]]]] = {}
        for param_name, param_info in params.items():
            if not isinstance(param_info, dict):
                continue
            original = param_info.get("original_value")
            if original is None:
                continue
            lookup.setdefault(str(original), []).append((str(param_name), param_info))
        return lookup

    def _maybe_parameterize_value(self, value: str) -> str:
        candidates = self._param_lookup.get(value) or []
        if not candidates:
            return repr(value)

        if len(candidates) == 1:
            param_name, param_info = candidates[0]
        else:
            cursor = self._param_cursors.get(value, 0)
            param_name, param_info = candidates[min(cursor, len(candidates) - 1)]
            self._param_cursors[value] = cursor + 1

        if param_info.get("sensitive"):
            return f"kwargs[{param_name!r}]"
        return f"kwargs.get({param_name!r}, {value!r})"

    def _render_embedded_ai_code_trace(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        previous_traces: List[RPAAcceptedTrace],
        used_output_keys: Dict[str, int],
    ) -> List[str]:
        key = self._allocate_output_key(trace, trace.output_key, used_output_keys) if trace.output_key else ""
        code = self._rewrite_dynamic_urls_in_code(
            (trace.ai_execution.code if trace.ai_execution else "").strip(),
            previous_traces,
        )
        lines = ["", f"    # trace {index}: {trace.description or 'AI operation'}"]
        for code_line in code.splitlines():
            lines.append(f"    {code_line}" if code_line.strip() else "")
        lines.append("    _result = await run(current_page, _results)")
        if key:
            lines.append(f"    _results[{key!r}] = _result")
        return lines

    def _render_dataflow_fill_trace(self, index: int, trace: RPAAcceptedTrace) -> List[str]:
        ref = trace.dataflow.selected_source_ref if trace.dataflow else None
        locator = self._best_locator(trace.dataflow.target_field.locator_candidates if trace.dataflow else [])
        lines = ["", f"    # trace {index}: dataflow fill {ref or ''}"]
        if not ref or not locator:
            lines.append("    # Unresolved dataflow fill skipped.")
            return lines
        lines.append(f"    _value = _resolve_result_ref(_results, {ref!r})")
        lines.append(f"    await {_locator_expression('current_page', locator)}.fill(str(_value))")
        return lines

    def _best_locator(self, candidates: List[Dict[str, Any]]) -> Dict[str, Any]:
        if not candidates:
            return {}
        selected = next((item for item in candidates if item.get("selected")), candidates[0])
        locator = selected.get("locator") if isinstance(selected, dict) else None
        return locator if isinstance(locator, dict) else selected

    def _dynamic_url_expression(self, url: str, previous_traces: List[RPAAcceptedTrace]) -> str:
        if not url:
            return ""
        for trace in reversed(previous_traces):
            result_expr = self._trace_result_url_expression(trace)
            output = trace.output if isinstance(trace.output, dict) else {}
            base = output.get("url") or output.get("value")
            if result_expr and isinstance(base, str) and base and url.startswith(base):
                suffix = url[len(base):]
                return f"str({result_expr}).rstrip('/') + {suffix!r}"
            observed_base = _repo_base_from_url(trace.after_page.url) if trace.after_page.url else ""
            if result_expr and observed_base and url.startswith(observed_base):
                suffix = url[len(observed_base):]
                return f"str({result_expr}).rstrip('/') + {suffix!r}"
        return ""

    def _previous_repo_url_expression(self, previous_traces: List[RPAAcceptedTrace]) -> str:
        for trace in reversed(previous_traces):
            result_expr = self._trace_result_url_expression(trace)
            if result_expr:
                return result_expr
        return ""

    def _trace_result_url_expression(self, trace: RPAAcceptedTrace) -> str:
        key = self._compiled_output_keys.get(id(trace), trace.output_key or "")
        if not key:
            return ""
        output = trace.output if isinstance(trace.output, dict) else {}
        if output.get("url"):
            return f"_resolve_result_ref(_results, {key + '.url'!r})"
        if output.get("value"):
            return f"_resolve_result_ref(_results, {key + '.value'!r})"
        instruction = f"{trace.user_instruction or ''} {trace.description or ''}".lower()
        if (
            trace.trace_type == RPATraceType.AI_OPERATION
            and (
                _looks_like_highest_star(instruction)
                or _looks_like_semantic_repo_selection(instruction, trace.output)
            )
        ):
            return f"_resolve_first_result_ref(_results, [{key + '.url'!r}, {key + '.value'!r}])"
        return ""

    def _manual_github_subpage_navigation(
        self,
        index: int,
        trace: RPAAcceptedTrace,
        previous_traces: List[RPAAcceptedTrace],
    ) -> List[str]:
        suffix = _manual_github_subpage_suffix(trace)
        if not suffix:
            return []
        repo_expr = self._previous_repo_url_expression(previous_traces)
        if not repo_expr:
            return []
        return [
            "",
            f"    # trace {index}: {trace.description or 'stable GitHub repository subpage navigation'}",
            f"    _repo_base = _github_repo_base(str({repo_expr}))",
            "    if not _repo_base:",
            "        raise RuntimeError('Could not resolve GitHub repository URL for recorded subpage navigation')",
            f"    _target_url = _repo_base + {suffix!r}",
            "    await current_page.goto(_target_url, wait_until='domcontentloaded')",
            "    await current_page.wait_for_load_state('domcontentloaded')",
        ]

    def _rewrite_dynamic_urls_in_code(self, code: str, previous_traces: List[RPAAcceptedTrace]) -> str:
        if not code or not previous_traces:
            return code

        def replace(match: re.Match[str]) -> str:
            url = match.group("url")
            dynamic = self._dynamic_url_expression(url, previous_traces)
            return dynamic or match.group(0)

        return re.sub(
            r"(?P<quote>['\"])(?P<url>https://github\.com/[^'\"]+)(?P=quote)",
            replace,
            code,
        )

    def _allocate_output_key(
        self,
        trace: RPAAcceptedTrace,
        raw_key: Optional[str],
        used_output_keys: Dict[str, int],
    ) -> str:
        key = str(raw_key or "").strip()
        if not key:
            return ""
        count = used_output_keys.get(key, 0) + 1
        used_output_keys[key] = count
        allocated = key if count == 1 else f"{key}_{count}"
        self._compiled_output_keys[id(trace)] = allocated
        return allocated


def _locator_expression(scope: str, locator: Dict[str, Any]) -> str:
    method = locator.get("method")
    if method == "role" or (method is None and locator.get("role")):
        role = locator.get("role", "button")
        name = locator.get("name")
        exact = locator.get("exact")
        args = [repr(role)]
        kwargs = []
        if name:
            kwargs.append(f"name={name!r}")
        if exact is not None:
            kwargs.append(f"exact={bool(exact)!r}")
        return f"{scope}.get_by_role({', '.join(args + kwargs)})"
    if method == "text":
        value = locator.get("value", "")
        exact = locator.get("exact")
        suffix = f", exact={bool(exact)!r}" if exact is not None else ""
        return f"{scope}.get_by_text({value!r}{suffix})"
    if method == "testid":
        return f"{scope}.get_by_test_id({locator.get('value', '')!r})"
    if method == "label":
        return f"{scope}.get_by_label({locator.get('value', '')!r})"
    if method == "placeholder":
        return f"{scope}.get_by_placeholder({locator.get('value', '')!r})"
    if method == "nested":
        parent = _locator_expression(scope, locator.get("parent") or {})
        return _locator_expression(parent, locator.get("child") or {})
    if method == "nth":
        base = _locator_expression(scope, locator.get("locator") or locator.get("base") or {"method": "css", "value": "body"})
        return f"{base}.nth({int(locator.get('index') or 0)})"
    return f"{scope}.locator({locator.get('value', 'body')!r}).first"


def _looks_like_highest_star(text: str) -> bool:
    return any(pattern in text for pattern in ("highest star", "most stars", "star count", "star数量最多", "start数量最多", "最多的项目"))


def _looks_like_pr_extraction(text: str, output: Any) -> bool:
    return (
        ("pr" in text or "pull request" in text or "pull requests" in text)
        and ("title" in text or "标题" in text)
        and ("creator" in text or "author" in text or "创建人" in text)
    ) or (isinstance(output, list) and output and isinstance(output[0], dict) and "title" in output[0])


def _looks_like_semantic_repo_selection(text: str, output: Any) -> bool:
    return (
        ("related" in text or "相关" in text or "semantic" in text)
        and ("repo" in text or "project" in text or "项目" in text)
        and isinstance(output, dict)
        and bool(output.get("url") or output.get("value"))
    )


def _extract_semantic_query(text: str) -> str:
    for pattern in (r"related to\s+([a-zA-Z0-9_+#.-]+)", r"和\s*([a-zA-Z0-9_+#.-]+)\s*最相关", r"most related to\s+([a-zA-Z0-9_+#.-]+)"):
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    return text


def _looks_like_highest_star(text: str) -> bool:
    return any(
        pattern in text
        for pattern in (
            "highest star",
            "most stars",
            "star count",
            "stars最多",
            "star数最多",
            "start数量最多",
            "star数量最多",
            "最多星",
            "最多的项目",
        )
    )


def _looks_like_pr_extraction(text: str, output: Any) -> bool:
    return (
        ("pr" in text or "pull request" in text or "pull requests" in text)
        and ("title" in text or "标题" in text)
        and ("creator" in text or "author" in text or "创建人" in text)
    ) or (isinstance(output, list) and output and isinstance(output[0], dict) and "title" in output[0])


def _looks_like_semantic_repo_selection(text: str, output: Any) -> bool:
    return (
        ("related" in text or "相关" in text or "semantic" in text)
        and ("repo" in text or "repository" in text or "project" in text or "项目" in text)
    )


def _extract_semantic_query(text: str) -> str:
    for pattern in (
        r"related to\s+([a-zA-Z0-9_+#.-]+)",
        r"和\s*([a-zA-Z0-9_+#.-]+)\s*最相关",
        r"most related to\s+([a-zA-Z0-9_+#.-]+)",
    ):
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    return text


def _extract_requested_page_count(text: str) -> int:
    normalized = str(text or "").lower()
    chinese_numbers = {"一": 1, "两": 2, "二": 2, "三": 3, "四": 4, "五": 5}
    match = re.search(r"前\s*(\d+)\s*页", normalized)
    if match:
        return max(int(match.group(1)), 1)
    for word, value in chinese_numbers.items():
        if f"前{word}页" in normalized or f"{word}页" in normalized:
            return value
    match = re.search(r"(?:first|top)\s+(\d+)\s+pages?", normalized)
    if match:
        return max(int(match.group(1)), 1)
    match = re.search(r"(\d+)\s+pages?", normalized)
    if match:
        return max(int(match.group(1)), 1)
    if "two pages" in normalized:
        return 2
    return 1


def _extract_record_limit(text: str) -> Optional[int]:
    normalized = str(text or "").lower()
    match = re.search(r"前\s*(\d+)\s*(?:个|条)", normalized)
    if match:
        return max(int(match.group(1)), 1)
    match = re.search(r"(?:first|top)\s+(\d+)\s+(?:prs?|pull requests?|records?|items?)", normalized)
    if match:
        return max(int(match.group(1)), 1)
    if any(pattern in normalized for pattern in ("前十个", "前十条")):
        return 10
    return None


def _repo_base_from_url(url: str) -> str:
    text = str(url or "").split("?", 1)[0].rstrip("/")
    match = re.match(r"(https://github\.com/[^/]+/[^/]+)", text)
    return match.group(1) if match else ""


def _trace_upload_source(trace: RPAAcceptedTrace) -> Dict[str, Any]:
    signals = trace.signals if isinstance(trace.signals, dict) else {}
    source = signals.get("upload_source")
    if isinstance(source, dict):
        return source
    return default_upload_source_from_staging(signals, fallback_filename=str(trace.value or "upload.bin"))


def _trace_download_signal(trace: RPAAcceptedTrace) -> Dict[str, Any]:
    signals = trace.signals if isinstance(trace.signals, dict) else {}
    download = signals.get("download")
    return download if isinstance(download, dict) else {}


def _is_github_repo_url(url: str) -> bool:
    return bool(_repo_base_from_url(url))


def _manual_github_subpage_suffix(trace: RPAAcceptedTrace) -> str:
    text_parts = [trace.description or "", trace.action or "", trace.after_page.url or ""]
    for candidate in trace.locator_candidates or []:
        locator = candidate.get("locator") if isinstance(candidate, dict) else candidate
        if isinstance(locator, dict):
            text_parts.extend(str(locator.get(key) or "") for key in ("name", "value", "role"))
    text = " ".join(text_parts).lower()
    if "/pulls" in text or "pull request" in text or "pull requests" in text or re.search(r"\bprs?\b", text):
        return "/pulls?q=is%3Apr"
    if "/issues" in text or "issues" in text or "issue" in text:
        return "/issues"
    if "/actions" in text or "actions" in text:
        return "/actions"
    if "/releases" in text or "releases" in text:
        return "/releases"
    return ""


def _runner_template(is_local: bool) -> str:
    if is_local:
        return '''\
import asyncio
import json as _json
import re
import sys
import time
from playwright.async_api import async_playwright

{execute_skill_func}


async def main():
    kwargs = {{}}
    for arg in sys.argv[1:]:
        if arg.startswith("--") and "=" in arg:
            k, v = arg[2:].split("=", 1)
            kwargs[k] = v
    pw = await async_playwright().start()
    browser = await pw.chromium.launch(**{launch_kwargs})
    context = await browser.new_context(**{context_kwargs})
    page = await context.new_page()
    page.set_default_timeout(60000)
    page.set_default_navigation_timeout(60000)
    try:
        result = await execute_skill(page, **kwargs)
        if result:
            print("SKILL_DATA:" + _json.dumps(result, ensure_ascii=False, default=str))
        print("SKILL_SUCCESS")
    except Exception as exc:
        print(f"SKILL_ERROR: {{exc}}", file=sys.stderr)
        sys.exit(1)
    finally:
        await context.close()
        await browser.close()
        await pw.stop()


if __name__ == "__main__":
    asyncio.run(main())
'''
    return '''\
import asyncio
import json as _json
import re
import sys
import time
import httpx
from playwright.async_api import async_playwright


async def _get_cdp_url() -> str:
    async with httpx.AsyncClient(timeout=10) as client:
        resp = await client.get("http://127.0.0.1:8080/v1/browser/info")
        resp.raise_for_status()
        return resp.json()["data"]["cdp_url"]


{execute_skill_func}


async def main():
    kwargs = {{}}
    for arg in sys.argv[1:]:
        if arg.startswith("--") and "=" in arg:
            k, v = arg[2:].split("=", 1)
            kwargs[k] = v
    cdp_url = await _get_cdp_url()
    pw = await async_playwright().start()
    browser = await pw.chromium.connect_over_cdp(cdp_url)
    context = await browser.new_context(**{context_kwargs})
    page = await context.new_page()
    page.set_default_timeout(60000)
    page.set_default_navigation_timeout(60000)
    try:
        result = await execute_skill(page, **kwargs)
        if result:
            print("SKILL_DATA:" + _json.dumps(result, ensure_ascii=False, default=str))
        print("SKILL_SUCCESS")
    except Exception as exc:
        print(f"SKILL_ERROR: {{exc}}", file=sys.stderr)
        sys.exit(1)
    finally:
        await context.close()
        await pw.stop()


if __name__ == "__main__":
    asyncio.run(main())
'''
