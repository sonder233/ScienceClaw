from __future__ import annotations

import asyncio
import json
import re
import sys
import uuid
from pathlib import Path
from typing import Any, Dict, Optional

from langchain_core.messages import HumanMessage, SystemMessage

from backend.config import settings
from backend.deepagent.engine import get_llm_model
from backend.rpa.upload_source import safe_asset_filename, safe_name_stem


TRANSFORM_TIMEOUT_S = 60.0


def transform_result_key(filename: str) -> str:
    return f"transform_{safe_name_stem(filename, fallback='file')}"


def transform_workspace_dir(session_id: str, transform_id: str) -> Path:
    return Path(settings.workspace_dir) / "rpa_transforms" / session_id / transform_id


def extract_text(response: Any) -> str:
    content = getattr(response, "content", response)
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for item in content:
            if isinstance(item, dict):
                parts.append(str(item.get("text") or item.get("content") or ""))
            else:
                parts.append(str(item))
        return "\n".join(part for part in parts if part)
    return str(content or "")


def extract_python_code(text: str) -> str:
    match = re.search(r"```(?:python)?\s*(.*?)```", text or "", re.DOTALL | re.IGNORECASE)
    return (match.group(1) if match else text).strip()


def ensure_transform_script(code: str) -> str:
    script = extract_python_code(code)
    if "def transform_file(" not in script:
        raise ValueError("Generated transform script must define transform_file(input_file, output_file, instruction='')")
    wrapper = '''


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--input-file", required=True)
    parser.add_argument("--output-file", required=True)
    parser.add_argument("--instruction", default="")
    args = parser.parse_args()
    transform_file(args.input_file, args.output_file, args.instruction)
'''
    if "if __name__ ==" not in script:
        script = script.rstrip() + wrapper
    return script


def fallback_transform_script() -> str:
    return ensure_transform_script(
        r'''
from __future__ import annotations

import csv
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _copy_worksheet(src, dst):
    for row in src.iter_rows():
        dst.append([cell.value for cell in row])


def transform_file(input_file, output_file, instruction=""):
    input_path = Path(input_file)
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = input_path.suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(input_path)
        wb.save(output_path)
        return str(output_path)

    if suffix in {".csv", ".tsv", ".txt"}:
        delimiter = "\t" if suffix in {".tsv", ".txt"} else ","
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        with input_path.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.reader(fh, delimiter=delimiter):
                ws.append(row)
        wb.save(output_path)
        return str(output_path)

    shutil.copy2(input_path, output_path)
    return str(output_path)
'''
    )


def preview_file(input_path: Path, max_rows: int = 12, max_chars: int = 6000) -> Dict[str, Any]:
    suffix = input_path.suffix.lower()
    preview: Dict[str, Any] = {
        "filename": input_path.name,
        "suffix": suffix,
        "size": input_path.stat().st_size if input_path.exists() else 0,
    }
    try:
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            wb = load_workbook(input_path, read_only=True, data_only=True)
            sheets = []
            for ws in wb.worksheets[:3]:
                rows = []
                for index, row in enumerate(ws.iter_rows(values_only=True)):
                    if index >= max_rows:
                        break
                    rows.append([cell for cell in row])
                sheets.append({"name": ws.title, "rows": rows})
            wb.close()
            preview["sheets"] = sheets
        elif suffix in {".csv", ".tsv", ".txt"}:
            text = input_path.read_text(encoding="utf-8-sig", errors="replace")
            preview["text"] = text[:max_chars]
        else:
            preview["note"] = "Binary or unsupported preview type"
    except Exception as exc:
        preview["preview_error"] = str(exc)
    return preview


class FileTransformService:
    async def compile_script(
        self,
        *,
        instruction: str,
        input_path: Path,
        output_filename: str,
        model_config: Optional[Dict[str, Any]] = None,
    ) -> str:
        if not instruction.strip():
            return fallback_transform_script()

        prompt_payload = {
            "instruction": instruction,
            "input_preview": preview_file(input_path),
            "output_filename": output_filename,
        }
        system_prompt = (
            "You generate deterministic Python file transformation code for an RPA workflow. "
            "Return only Python code. The code must define exactly one callable named "
            "transform_file(input_file, output_file, instruction=''). Use only Python standard "
            "library and openpyxl. Do not use pandas. The function must create a valid .xlsx "
            "file at output_file, preserve source data unless the instruction asks to reshape it, "
            "and raise a clear exception if the input cannot be transformed."
        )
        model = get_llm_model(config=model_config, streaming=False)
        response = await model.ainvoke(
            [
                SystemMessage(content=system_prompt),
                HumanMessage(content=json.dumps(prompt_payload, ensure_ascii=False, default=str)),
            ]
        )
        return ensure_transform_script(extract_text(response))

    async def execute_script(
        self,
        *,
        script_path: Path,
        input_path: Path,
        output_path: Path,
        instruction: str,
        timeout: float = TRANSFORM_TIMEOUT_S,
    ) -> Dict[str, Any]:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        proc = await asyncio.create_subprocess_exec(
            sys.executable,
            str(script_path),
            "--input-file",
            str(input_path),
            "--output-file",
            str(output_path),
            "--instruction",
            instruction,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        try:
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        except asyncio.TimeoutError as exc:
            proc.kill()
            await proc.communicate()
            raise TimeoutError(f"File transform exceeded {timeout:.0f}s") from exc

        if proc.returncode != 0:
            detail = stderr.decode("utf-8", errors="replace").strip() or stdout.decode("utf-8", errors="replace").strip()
            raise RuntimeError(detail or f"File transform failed with exit code {proc.returncode}")

        if not output_path.exists():
            raise RuntimeError("File transform did not create the expected output file")

        try:
            from openpyxl import load_workbook

            wb = load_workbook(output_path, read_only=True, data_only=True)
            sheet_count = len(wb.sheetnames)
            first_sheet = wb.sheetnames[0] if wb.sheetnames else ""
            wb.close()
        except Exception as exc:
            raise RuntimeError(f"Generated file is not a valid .xlsx workbook: {exc}") from exc

        return {
            "filename": output_path.name,
            "path": str(output_path),
            "size": output_path.stat().st_size,
            "sheet_count": sheet_count,
            "first_sheet": first_sheet,
        }

    async def create_transform(
        self,
        *,
        session_id: str,
        input_path: Path,
        instruction: str,
        output_filename: str,
        model_config: Optional[Dict[str, Any]] = None,
        output_key: Optional[str] = None,
    ) -> Dict[str, Any]:
        transform_id = f"tf_{uuid.uuid4().hex}"
        safe_output = safe_asset_filename(output_filename or f"{input_path.stem}_converted.xlsx")
        if not safe_output.lower().endswith(".xlsx"):
            safe_output = f"{Path(safe_output).stem or 'converted'}.xlsx"
        result_key = output_key or transform_result_key(safe_output)
        work_dir = transform_workspace_dir(session_id, transform_id)
        script_path = work_dir / "transform.py"
        output_path = work_dir / safe_output

        work_dir.mkdir(parents=True, exist_ok=True)
        script = await self.compile_script(
            instruction=instruction,
            input_path=input_path,
            output_filename=safe_output,
            model_config=model_config,
        )

        script_path.write_text(script, encoding="utf-8")
        result = await self.execute_script(
            script_path=script_path,
            input_path=input_path,
            output_path=output_path,
            instruction=instruction,
        )
        spec = {
            "transform_id": transform_id,
            "instruction": instruction,
            "input_path": str(input_path),
            "output_filename": safe_output,
            "output_result_key": result_key,
            "script_path": str(script_path),
            "output_path": str(output_path),
            "code": script,
        }
        (work_dir / "transform_spec.json").write_text(
            json.dumps(spec, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return {
            **spec,
            "result": result,
        }
